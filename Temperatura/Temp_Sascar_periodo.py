# -*- coding: utf-8 -*-
# Extrair temperaturas de PDF (Sascar) -> Excel + gráfico com faixas
# Fabio Cordeiro - versão 1.1 (Com filtro de data no gráfico)

import re
import os
import math
import datetime as dt
from pathlib import Path

import pdfplumber
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.dates import DateFormatter
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage

# =========================
# CONFIGURAÇÕES
# =========================
# Caminho do PDF (ajuste se necessário)
PDF_PATH = r"C:\\Fabio\\Desenvolvimento\\Varejo\\Temperatura\\Temp_Sascar.pdf"
# Saídas
OUT_XLSX = r"C:\\Fabio\\Desenvolvimento\\Varejo\\Temperatura\\Temperatura_Extraida.xlsx"
OUT_PNG  = r"C:\\Fabio\\Desenvolvimento\\Varejo\\Temperatura\\Grafico_Temperatura.png"

# Qual coluna priorizar como "temperatura apurada"
APURADA_PRIORIDADE = ["TS", "T1", "T2", "T3"]

# =========================
# CONFIGURAÇÃO DO FILTRO DE DATA (GRÁFICO)
# =========================
# Definindo o range solicitado: 07/02/2026 00:01 até 09/02/2026 12:00
FILTRO_INICIO = dt.datetime(2026, 2, 7, 0, 1)
FILTRO_FIM    = dt.datetime(2026, 2, 9, 12, 0)

# =========================
# FUNÇÕES AUXILIARES
# =========================
RE_DATAHORA = re.compile(r"(\d{2}/\d{2}/\d{4}\s+\d{2}:\d{2}:\d{2})")

def _parse_float_cell(txt: str):
    """Converte a célula de temperatura para float ou None."""
    if not txt:
        return None
    s = str(txt).strip().upper()
    if s in {"", "N/A", "NA"}:
        return None
    s = s.replace("—", "-").replace("−", "-")
    s = re.sub(r"^-{2,}", "-", s)
    s = re.sub(r"[^0-9\-\.,]+", "", s)
    if s == "" or s == "-":
        return None
    s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        m = re.search(r"-?\d+(?:\.\d+)?", s)
        return float(m.group(0)) if m else None

def _parse_datetime(br_str: str):
    return dt.datetime.strptime(br_str, "%d/%m/%Y %H:%M:%S")

def escolher_temperatura_apurada(row) -> float | None:
    for col in APURADA_PRIORIDADE:
        v = row.get(col)
        if v is not None and not (isinstance(v, float) and math.isnan(v)):
            return v
    return None

# =========================
# EXTRAÇÃO DO PDF
# =========================
def extrair_do_pdf(pdf_path: str) -> pd.DataFrame:
    registros = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text(x_tolerance=2, y_tolerance=2) or ""
            if not text.strip():
                continue
            for raw_line in text.splitlines():
                line = " ".join(raw_line.split())
                m = RE_DATAHORA.search(line)
                if not m:
                    continue
                dh_str = m.group(1)
                try:
                    dh = _parse_datetime(dh_str)
                except Exception:
                    continue

                vals = {}
                # Busca explícita (T1 -10, etc)
                for col in ["T1", "T2", "T3", "TS"]:
                    mm = re.search(rf"\b{col}\b\s*([\-–—]?\d+[,\.]?\d*)", line, flags=re.IGNORECASE)
                    if mm:
                        vals[col] = _parse_float_cell(mm.group(1))

                # Fallback posicional
                if not vals:
                    candidates = re.findall(r"[\-–—]?\d{1,2}[,\.]?\d*", line)
                    tail = [candidates[i] for i in range(max(0, len(candidates)-4), len(candidates))]
                    tail = [_parse_float_cell(t) for t in tail]
                    tail = [t for t in tail if t is not None]
                    if len(tail) >= 1:
                        mapping_seq = {
                            1: ["TS"],
                            2: ["T1", "TS"],
                            3: ["T1", "T2", "TS"],
                            4: ["T1", "T2", "T3", "TS"],
                        }
                        cols = mapping_seq.get(min(4, len(tail)))
                        if cols:
                            for c, v in zip(cols, tail[-len(cols):]):
                                vals.setdefault(c, v)

                if not vals:
                    continue

                rec = {"DataHora": dh,
                       "T1": vals.get("T1"),
                       "T2": vals.get("T2"),
                       "T3": vals.get("T3"),
                       "TS": vals.get("TS")}
                registros.append(rec)

    df = pd.DataFrame(registros).sort_values("DataHora").reset_index(drop=True)
    df["Temperatura_Apurada"] = df.apply(escolher_temperatura_apurada, axis=1)
    return df

# =========================
# GRÁFICO (PNG) + EXCEL
# =========================
def criar_grafico_png(df: pd.DataFrame, out_png: str):
    if df.empty:
        print("⚠️ Aviso: Não há dados para plotar no período selecionado.")
        return

    fig, ax = plt.subplots(figsize=(10, 4))
    ax.plot(df["DataHora"], df["Temperatura_Apurada"], linewidth=1.5, marker='.', markersize=4)

    # Faixas
    ax.axhspan(-40, -12, alpha=0.2, color='blue', label='≤ -12°C')
    ax.axhspan(0, 5, alpha=0.2, color='green', label='0 a 5°C')

    ax.set_xlabel("Data/Hora")
    ax.set_ylabel("Temperatura (°C)")
    
    # Título dinâmico com as datas do DataFrame filtrado
    min_date = df["DataHora"].min().strftime("%d/%m %H:%M")
    max_date = df["DataHora"].max().strftime("%d/%m %H:%M")
    ax.set_title(f"Temperatura Apurada ({min_date} a {max_date})")
    
    ax.grid(True, linewidth=0.4, linestyle="--", alpha=0.6)

    ax.xaxis.set_major_formatter(DateFormatter("%d/%m %H:%M"))
    fig.autofmt_xdate()

    plt.tight_layout()
    fig.savefig(out_png, dpi=150)
    plt.close(fig)

def salvar_excel_com_grafico(df: pd.DataFrame, out_xlsx: str, img_path: str):
    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Dados")

    wb = load_workbook(out_xlsx)
    ws = wb.create_sheet("Gráfico")
    if Path(img_path).exists():
        img = XLImage(img_path)
        img.anchor = "A1"
        ws.add_image(img)
    
    ws["A30"] = "Informações:"
    ws["A31"] = f"Gráfico gerado para o período: {FILTRO_INICIO} até {FILTRO_FIM}"
    ws["A32"] = "Faixas: ≤ -12 °C e 0 a 5 °C"
    wb.save(out_xlsx)

# =========================
# MAIN
# =========================
def main():
    pdf_file = PDF_PATH
    if not Path(pdf_file).exists():
        raise FileNotFoundError(f"PDF não encontrado: {pdf_file}")

    print("[1/3] Extraindo dados COMPLETOS do PDF...")
    df_completo = extrair_do_pdf(pdf_file)

    if df_completo.empty:
        raise RuntimeError("Nenhuma linha de temperatura foi reconhecida no PDF.")

    # ---------------------------------------------------------
    # APLICANDO O FILTRO SOLICITADO APENAS PARA O GRÁFICO
    # ---------------------------------------------------------
    print(f"[2/3] Filtrando gráfico para o período: {FILTRO_INICIO} até {FILTRO_FIM}...")
    
    mask = (df_completo["DataHora"] >= FILTRO_INICIO) & (df_completo["DataHora"] <= FILTRO_FIM)
    df_grafico = df_completo.loc[mask].copy()

    if df_grafico.empty:
        print(f"⚠️ ATENÇÃO: Nenhum dado encontrado entre {FILTRO_INICIO} e {FILTRO_FIM}.")
        print("   -> Gerando gráfico com TODOS os dados como fallback.")
        criar_grafico_png(df_completo, OUT_PNG)
    else:
        criar_grafico_png(df_grafico, OUT_PNG)
        print(f"   -> Gráfico gerado com {len(df_grafico)} registros filtrados.")

    print("[3/3] Salvando Excel (Dados completos + Gráfico)...")
    salvar_excel_com_grafico(df_completo, OUT_XLSX, OUT_PNG)

    print(f"Concluído.")
    print(f"Excel salvo em: {OUT_XLSX}")

if __name__ == "__main__":
    main()
# -*- coding: utf-8 -*-
# Extrair temperaturas de PDF (Sascar) -> Excel + gráfico com faixas
# Fabio Cordeiro - versão 3.0
# Script Linear de Extração de Temperaturas (Sascar)
# Ajustes: Escala de 5 em 5, Linha Azul Escuro, Congelado até -25.

import re
import os
import math
import datetime as dt
from pathlib import Path

import pdfplumber
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker  # Importado para controlar a escala de 5 em 5
from matplotlib.dates import DateFormatter
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

# ==========================================
# 1. CONFIGURAÇÕES E CAMINHOS
# ==========================================
PDF_PATH = r"C:\\Fabio\\Desenvolvimento\\Varejo\\Temperatura\\Temp_Sascar.pdf"
OUT_XLSX = r"C:\\Fabio\\Desenvolvimento\\Varejo\\Temperatura\\Temperatura_Extraida.xlsx"
OUT_PNG  = r"C:\\Fabio\\Desenvolvimento\\Varejo\\Temperatura\\Grafico_Temperatura.png"

# Placa do veículo (para exibir DENTRO do gráfico, centralizado)
PLACA_VEICULO = "Transportadora Luiz Mauro Hergert - Placa: FQH1H55 - FQS1A24"  # ajuste aqui

# Período de Filtro exemplo
#DATA_INICIO = dt.datetime(2026, 2, 7, 0, 1)   # 07/02/2026 00:01
#DATA_FIM    = dt.datetime(2026, 2, 9, 12, 0)  # 09/02/2026 12:00
DATA_INICIO = dt.datetime(2026, 3, 15, 20, 1)
DATA_FIM    = dt.datetime(2026, 3, 17, 11, 40)

# Prioridade de colunas
APURADA_PRIORIDADE = ["TS", "T1", "T2", "T3"]

# Regex para data
RE_DATAHORA = re.compile(r"(\d{2}/\d{2}/\d{4}\s+\d{2}:\d{2}:\d{2})")

print(">>> Iniciando processamento linear...")

# ==========================================
# 2. EXTRAÇÃO DE DADOS
# ==========================================
registros = []

if not os.path.exists(PDF_PATH):
    raise FileNotFoundError(f"PDF não encontrado: {PDF_PATH}")

print("[1/4] Lendo PDF...")

with pdfplumber.open(PDF_PATH) as pdf:
    for page in pdf.pages:
        text = page.extract_text(x_tolerance=2, y_tolerance=2) or ""
        if not text.strip():
            continue

        for raw_line in text.splitlines():
            line = " ".join(raw_line.split())

            m = RE_DATAHORA.search(line)
            if not m:
                continue

            dh_str = m.group(1)
            try:
                dh = dt.datetime.strptime(dh_str, "%d/%m/%Y %H:%M:%S")
            except:
                continue

            vals = {}

            # Parsing: Busca rótulos explícitos
            for col in ["T1", "T2", "T3", "TS"]:
                mm = re.search(rf"\b{col}\b\s*([\-–—]?\d+[,\.]?\d*)", line, flags=re.IGNORECASE)
                if mm:
                    txt = mm.group(1).strip().upper().replace("—", "-").replace("−", "-")
                    txt = re.sub(r"[^0-9\-\.,]+", "", txt)
                    txt = txt.replace(",", ".")
                    try:
                        vals[col] = float(txt)
                    except:
                        pass

            # Parsing: Busca sequência numérica (fallback)
            if not vals:
                candidates = re.findall(r"[\-–—]?\d{1,2}[,\.]?\d*", line)
                tail_raw = [candidates[i] for i in range(max(0, len(candidates)-4), len(candidates))]
                tail_float = []
                for t_str in tail_raw:
                    t_clean = t_str.replace("—", "-").replace("−", "-").replace(",", ".")
                    t_clean = re.sub(r"[^0-9\-\.]+", "", t_clean)
                    try:
                        tail_float.append(float(t_clean))
                    except:
                        pass

                if len(tail_float) >= 1:
                    mapping = {
                        1: ["TS"],
                        2: ["T1", "TS"],
                        3: ["T1", "T2", "TS"],
                        4: ["T1", "T2", "T3", "TS"]
                    }
                    cols_to_map = mapping.get(min(4, len(tail_float)))
                    if cols_to_map:
                        for c, v in zip(cols_to_map, tail_float[-len(cols_to_map):]):
                            vals.setdefault(c, v)

            if vals:
                rec = {
                    "DataHora": dh,
                    "T1": vals.get("T1"),
                    "T2": vals.get("T2"),
                    "T3": vals.get("T3"),
                    "TS": vals.get("TS")
                }
                registros.append(rec)

# Cria DataFrame
df = pd.DataFrame(registros)

if df.empty:
    raise RuntimeError("Nenhum dado encontrado no PDF.")

if "DataHora" not in df.columns:
    raise KeyError(f"Coluna 'DataHora' não encontrada. Colunas disponíveis: {list(df.columns)}")

df = df.sort_values("DataHora").reset_index(drop=True)

# Define Temperatura Apurada
def get_apurada_inline(row):
    for c in APURADA_PRIORIDADE:
        v = row.get(c)
        if pd.notna(v):
            return v
    return None

df["Temperatura_Apurada"] = df.apply(get_apurada_inline, axis=1)

print(f"[2/4] Filtrando dados ({DATA_INICIO} a {DATA_FIM})...")

# Filtro de Data
mask = (df["DataHora"] >= DATA_INICIO) & (df["DataHora"] <= DATA_FIM)
df = df.loc[mask].copy()

if df.empty:
    print("Aviso: O filtro de data removeu todos os registros.")
else:
    print(f"   -> Registros restantes: {len(df)}")

# ==========================================
# 3. GERAÇÃO DO GRÁFICO (Ajustado)
# ==========================================
if not df.empty:
    print("[3/4] Gerando gráfico com faixas e nova escala...")

    fig, ax = plt.subplots(figsize=(12, 6))

    # --- ALTERAÇÃO 1: Cor da linha (Azul Escuro) ---
    ax.plot(df["DataHora"], df["Temperatura_Apurada"], linewidth=2, color="#6E3AB8", label="Temp. Apurada")

    # --- FAIXAS ---
    # Refrigerado: 0 a 4
    ax.axhspan(0, 4, color="#08DB08", alpha=0.2, label="Refrigerado (0°C a 4°C)")

    # --- ALTERAÇÃO 2: Faixa Congelado (-12 a -25) ---
    # Obs: no axhspan a ordem dos parâmetros (min, max) não interfere, mas visualmente é de -25 a -12
    ax.axhspan(-25, -18, color="#4FADE4", alpha=0.4, label="Congelado (-18°C ou mais frio)")

    # --- TEXTOS NO CENTRO DAS FAIXAS (não altera a legenda) ---
    x0 = df["DataHora"].min()
    x1 = df["DataHora"].max()
    x_center = x0 + (x1 - x0) / 2

    ax.text(
        x_center, 2, "Área refrigerada",   # centro entre 0 e 4
        ha="center", va="center",
        fontsize=11, fontweight="bold",
        color="#0B6B0B",
        bbox=dict(facecolor="white", edgecolor="none", alpha=0.6, pad=2)
    )

    ax.text(
        x_center, -18.5, "Área congelada", # centro entre -25 e -18
        ha="center", va="center",
        fontsize=11, fontweight="bold",
        color="#0B3E66",
        bbox=dict(facecolor="white", edgecolor="none", alpha=0.6, pad=2)
    )

    # Configurações do gráfico
    ax.set_xlabel("Data/Hora")
    ax.set_ylabel("Temperatura (°C)")

    min_date_str = df["DataHora"].min().strftime("%d/%m %H:%M")
    max_date_str = df["DataHora"].max().strftime("%d/%m %H:%M")
    # Informação na parte superior do gráfico (mantida)
    ax.set_title(f"Monitoramento de Temperatura ({min_date_str} até {max_date_str})")

    # --- PLACA DENTRO DA ÁREA DO GRÁFICO (CENTRO, TOPO) ---
    # Usando coordenadas do eixo (0..1), fica sempre no mesmo lugar visual.
    ax.text(
        0.5, 0.92, f"{PLACA_VEICULO}",
        transform=ax.transAxes,
        ha="center", va="center",
        fontsize=12, fontweight="bold",
        color="black",
        bbox=dict(facecolor="white", edgecolor="none", alpha=0.7, pad=2)
    )

    ax.grid(True, linewidth=0.4, linestyle="--", alpha=0.6)

    # --- ALTERAÇÃO 3: Escala de 5 em 5 graus ---
    ax.yaxis.set_major_locator(ticker.MultipleLocator(5))

    # Formatação do eixo X
    ax.xaxis.set_major_formatter(DateFormatter("%d/%m %H:%M"))
    fig.autofmt_xdate()

    # Legenda
    ax.legend(loc='best', frameon=True, shadow=True)

    plt.tight_layout()
    fig.savefig(OUT_PNG, dpi=150)
    plt.close(fig)
    print(f"   -> Gráfico salvo em: {OUT_PNG}")

# ==========================================
# 4. SALVAR EXCEL
# ==========================================
if not df.empty:
    print("[4/4] Salvando Excel final...")

    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Dados")

    wb = load_workbook(OUT_XLSX)
    ws = wb.create_sheet("Gráfico")

    if os.path.exists(OUT_PNG):
        img = XLImage(OUT_PNG)
        img.anchor = "A1"
        ws.add_image(img)

    ws["A30"] = "Parâmetros utilizados:"
    ws["A31"] = f"Filtro Início: {DATA_INICIO}"
    ws["A32"] = f"Filtro Fim: {DATA_FIM}"
    ws["A34"] = "Legenda das Faixas:"
    ws["A35"] = "Refrigerado: 0°C a 4°C"
    ws["A36"] = "Congelado: -12°C a -25°C"  # Atualizado texto no Excel

    wb.save(OUT_XLSX)
    print(f"   -> Excel salvo em: {OUT_XLSX}")

print("\n>>> Processo concluído com sucesso.")
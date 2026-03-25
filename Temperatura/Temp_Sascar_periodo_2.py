# -*- coding: utf-8 -*-
# Script Linear de Extração de Temperaturas (Sascar)
# Sem funções, execução direta.
# Versão com faixas: Refrigerado (0 a 4) e Congelado (-12 a -35)

import re
import os
import math
import datetime as dt
from pathlib import Path

import pdfplumber
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.dates import DateFormatter
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

# ==========================================
# 1. CONFIGURAÇÕES E CAMINHOS
# ==========================================
PDF_PATH = r"C:\\Fabio\\Desenvolvimento\\Varejo\\Temperatura\\Temp_Sascar.pdf"
OUT_XLSX = r"C:\\Fabio\\Desenvolvimento\\Varejo\\Temperatura\\Temperatura_Extraida.xlsx"
OUT_PNG  = r"C:\\Fabio\\Desenvolvimento\\Varejo\\Temperatura\\Grafico_Temperatura.png"

# Período de Filtro
DATA_INICIO = dt.datetime(2026, 2, 7, 0, 1)   # 07/02/2026 00:01
DATA_FIM    = dt.datetime(2026, 2, 9, 12, 0)  # 09/02/2026 12:00

# Prioridade de colunas
APURADA_PRIORIDADE = ["TS", "T1", "T2", "T3"]

# Regex para data
RE_DATAHORA = re.compile(r"(\d{2}/\d{2}/\d{4}\s+\d{2}:\d{2}:\d{2})")

print(">>> Iniciando processamento linear...")

# ==========================================
# 2. EXTRAÇÃO DE DADOS (Lógica Linear)
# ==========================================
registros = []

if not os.path.exists(PDF_PATH):
    raise FileNotFoundError(f"PDF não encontrado: {PDF_PATH}")

print("[1/4] Lendo PDF...")

with pdfplumber.open(PDF_PATH) as pdf:
    for page in pdf.pages:
        # Extrai texto
        text = page.extract_text(x_tolerance=2, y_tolerance=2) or ""
        if not text.strip():
            continue

        # Processa linha a linha
        for raw_line in text.splitlines():
            line = " ".join(raw_line.split()) # normaliza espaços
            
            # Busca Data/Hora
            m = RE_DATAHORA.search(line)
            if not m:
                continue
            
            dh_str = m.group(1)
            try:
                dh = dt.datetime.strptime(dh_str, "%d/%m/%Y %H:%M:%S")
            except:
                continue

            # Dicionário temporário para as temperaturas desta linha
            vals = {}
            
            # --- Lógica de Parsing Inline (substituindo função _parse_float_cell) ---
            # Método A: Busca rótulos explícitos T1, T2, etc.
            for col in ["T1", "T2", "T3", "TS"]:
                mm = re.search(rf"\b{col}\b\s*([\-–—]?\d+[,\.]?\d*)", line, flags=re.IGNORECASE)
                if mm:
                    txt = mm.group(1).strip().upper().replace("—", "-").replace("−", "-")
                    # Remove caracteres estranhos
                    txt = re.sub(r"[^0-9\-\.,]+", "", txt)
                    txt = txt.replace(",", ".")
                    try:
                        vals[col] = float(txt)
                    except:
                        pass

            # Método B: Se falhar, busca sequência de números no fim da linha
            if not vals:
                candidates = re.findall(r"[\-–—]?\d{1,2}[,\.]?\d*", line)
                # Pega os últimos 4 candidatos numéricos
                tail_raw = [candidates[i] for i in range(max(0, len(candidates)-4), len(candidates))]
                tail_float = []
                for t_str in tail_raw:
                    t_clean = t_str.replace("—", "-").replace("−", "-").replace(",", ".")
                    t_clean = re.sub(r"[^0-9\-\.]+", "", t_clean)
                    try:
                        tail_float.append(float(t_clean))
                    except:
                        pass
                
                # Mapeia baseada na quantidade encontrada
                if len(tail_float) >= 1:
                    mapping = {
                        1: ["TS"],
                        2: ["T1", "TS"],
                        3: ["T1", "T2", "TS"],
                        4: ["T1", "T2", "T3", "TS"]
                    }
                    cols_to_map = mapping.get(min(4, len(tail_float)))
                    if cols_to_map:
                        # Atribui valores do fim para o começo da lista mapeada
                        for c, v in zip(cols_to_map, tail_float[-len(cols_to_map):]):
                            vals.setdefault(c, v)

            # Se achou algo, adiciona à lista
            if vals:
                rec = {
                    "DataHora": dh,
                    "T1": vals.get("T1"),
                    "T2": vals.get("T2"),
                    "T3": vals.get("T3"),
                    "TS": vals.get("TS")
                }
                registros.append(rec)

# Cria DataFrame
df = pd.DataFrame(registros)
if df.empty:
    raise RuntimeError("Nenhum dado encontrado no PDF.")

# Ordena
df = df.sort_values("DataHora").reset_index(drop=True)

# Lógica inline para definir "Temperatura Apurada" (substituindo função)
# Prioridade: TS -> T1 -> T2 -> T3
def get_apurada_inline(row):
    for c in APURADA_PRIORIDADE:
        v = row.get(c)
        if pd.notna(v):
            return v
    return None

df["Temperatura_Apurada"] = df.apply(get_apurada_inline, axis=1)

print(f"[2/4] Filtrando dados ({DATA_INICIO} a {DATA_FIM})...")

# --- APLICAÇÃO DO FILTRO ---
mask = (df["DataHora"] >= DATA_INICIO) & (df["DataHora"] <= DATA_FIM)
df = df.loc[mask].copy()

if df.empty:
    print("Aviso: O filtro de data removeu todos os registros.")
else:
    print(f"   -> Registros restantes: {len(df)}")

# ==========================================
# 3. GERAÇÃO DO GRÁFICO (Linear)
# ==========================================
if not df.empty:
    print("[3/4] Gerando gráfico com faixas...")
    
    fig, ax = plt.subplots(figsize=(12, 6))
    
    # Plota a linha principal
    ax.plot(df["DataHora"], df["Temperatura_Apurada"], linewidth=2, color='black', label="Temp. Apurada")

    # --- FAIXAS SOLICITADAS ---
    # 1. Faixa Refrigerado: 0 a 4 graus
    ax.axhspan(0, 4, color='green', alpha=0.2, label="Refrigerado (0°C a 4°C)")
    
    # 2. Faixa Congelado: -12 a -35 graus (usamos o menor primeiro no python: -35 a -12)
    ax.axhspan(-35, -12, color='blue', alpha=0.2, label="Congelado (-12°C a -35°C)")

    # Configurações do gráfico
    ax.set_xlabel("Data/Hora")
    ax.set_ylabel("Temperatura (°C)")
    
    min_date_str = df["DataHora"].min().strftime("%d/%m %H:%M")
    max_date_str = df["DataHora"].max().strftime("%d/%m %H:%M")
    ax.set_title(f"Monitoramento de Temperatura ({min_date_str} até {max_date_str})")
    
    ax.grid(True, linewidth=0.4, linestyle="--", alpha=0.6)
    
    # Formatação do eixo X (Datas)
    ax.xaxis.set_major_formatter(DateFormatter("%d/%m %H:%M"))
    fig.autofmt_xdate()

    # --- LEGENDA ---
    # Posiciona a legenda para não ficar em cima da linha, se possível 'best'
    ax.legend(loc='best', frameon=True, shadow=True)

    plt.tight_layout()
    fig.savefig(OUT_PNG, dpi=150)
    plt.close(fig)
    print(f"   -> Gráfico salvo em: {OUT_PNG}")

# ==========================================
# 4. SALVAR EXCEL (Linear)
# ==========================================
if not df.empty:
    print("[4/4] Salvando Excel final...")
    
    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Dados")

    # Insere a imagem
    wb = load_workbook(OUT_XLSX)
    ws = wb.create_sheet("Gráfico")
    
    if os.path.exists(OUT_PNG):
        img = XLImage(OUT_PNG)
        img.anchor = "A1"
        ws.add_image(img)
    
    # Informações textuais no Excel
    ws["A30"] = "Parâmetros utilizados:"
    ws["A31"] = f"Filtro Início: {DATA_INICIO}"
    ws["A32"] = f"Filtro Fim: {DATA_FIM}"
    ws["A34"] = "Legenda das Faixas:"
    ws["A35"] = "Refrigerado: 0°C a 4°C"
    ws["A36"] = "Congelado: -12°C a -35°C"
    
    wb.save(OUT_XLSX)
    print(f"   -> Excel salvo em: {OUT_XLSX}")

print("\n>>> Processo concluído com sucesso.")
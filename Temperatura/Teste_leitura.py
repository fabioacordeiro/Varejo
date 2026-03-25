# -*- coding: utf-8 -*-
# Extrair temperaturas de PDF (Sascar) -> Excel + gráfico com faixas
# Fabio Cordeiro - versão 6.0 (Suporte Paisagem + Fallback OCR)

import re
import os
import datetime as dt
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
from matplotlib.dates import DateFormatter
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

# Bibliotecas de extração
import pdfplumber
from PyPDF2 import PdfReader

# Configuração de OCR (Caso as bibliotecas de texto falhem)
try:
    import pytesseract
    from pdf2image import convert_from_path
    OCR_DISPONIVEL = True
except ImportError:
    OCR_DISPONIVEL = False

# ==========================================
# 1. CONFIGURAÇÕES E CAMINHOS
# ==========================================
PDF_PATH = r"C:\\Fabio\\Desenvolvimento\\Varejo\\Temperatura\\Temp_sem_cabecalho.pdf"
OUT_XLSX = r"C:\\Fabio\\Desenvolvimento\\Varejo\\Temperatura\\Temperatura_Extraida_sem_cabecalho.xlsx"
OUT_PNG  = r"C:\\Fabio\Desenvolvimento\\Varejo\\Temperatura\\Grafico_Temperatura_arquivo_sem_cabecalho.png"

PLACA_VEICULO = "Transportadora Luiz Mauro Hergert - Placa: FQH1H55, FQS1A24"
DATA_INICIO = dt.datetime(2026, 3, 13, 0, 1)
DATA_FIM    = dt.datetime(2026, 3, 17, 23, 0)

DEBUG = True
APURADA_PRIORIDADE = ["TS", "T1", "T2", "T3"]

# Ajuste caminhos do Tesseract/Poppler se necessário (Windows)
TESSERACT_CMD = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
POPPLER_PATH = r"C:\poppler\Library\bin"

# ==========================================
# 2. REGEX DE CAPTURA
# ==========================================
RE_DATAHORA_ANTIGO = re.compile(r"(\d{2}/\d{2}/\d{4}\s+\d{2}:\d{2}:\d{2})")
RE_DATA_BLOCO = re.compile(r"^\d{2}[-/]\d{2}[-/]\d{4}$")
RE_HORA_BLOCO = re.compile(r"^\d{2}:\d{2}:\d{2}$")
RE_NUM = re.compile(r"^[\-–—−]?\d{1,2}(?:[\,\.]\d+)?$")

def log(msg):
    if DEBUG: print(msg)

# ==========================================
# 3. MÓDULO DE EXTRAÇÃO (COM SUPORTE A PAISAGEM)
# ==========================================
def extrair_texto_pdf(path):
    linhas = []
    print("[1/4] Tentando extrair texto (pdfplumber)...")
    
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            # x_tolerance maior ajuda em formatos paisagem onde as colunas são distantes
            text = page.extract_text(x_tolerance=3, y_tolerance=3)
            if text:
                for ln in text.splitlines():
                    clean_ln = " ".join(ln.split()).strip()
                    if clean_ln:
                        linhas.append({"texto": clean_ln, "pagina": page.page_number, "origem": "pdfplumber"})
    
    if not linhas:
        log("[DEBUG] pdfplumber falhou. Tentando PyPDF2...")
        reader = PdfReader(path)
        for i, page in enumerate(reader.pages):
            text = page.extract_text()
            if text:
                for ln in text.splitlines():
                    clean_ln = " ".join(ln.split()).strip()
                    linhas.append({"texto": clean_ln, "pagina": i+1, "origem": "PyPDF2"})

    if not linhas and OCR_DISPONIVEL:
        log("[DEBUG] Texto não detectado. Iniciando OCR (pode demorar)...")
        pytesseract.pytesseract.tesseract_cmd = TESSERACT_CMD
        images = convert_from_path(path, dpi=300, poppler_path=POPPLER_PATH)
        for i, img in enumerate(images):
            text = pytesseract.image_to_string(img, lang="por")
            for ln in text.splitlines():
                clean_ln = " ".join(ln.split()).strip()
                if clean_ln:
                    linhas.append({"texto": clean_ln, "pagina": i+1, "origem": "OCR"})
                    
    return linhas

# ==========================================
# 4. PARSERS (ANTIGO + NOVO)
# ==========================================
def processar_dados(linhas):
    registros = []
    i = 0
    while i < len(linhas):
        linha_atual = linhas[i]["texto"]
        
        # Tenta Formato Antigo (Data e Hora na mesma linha)
        m_antigo = RE_DATAHORA_ANTIGO.search(linha_atual)
        if m_antigo:
            dh = dt.datetime.strptime(m_antigo.group(1), "%d/%m/%Y %H:%M:%S")
            # Busca temperaturas na linha
            temps = re.findall(r"[\-–—−]?\d{1,2}[,\.]\d*", linha_atual)
            registros.append({"DataHora": dh, "TS": float(temps[-1].replace(',','.')) if temps else None, "Formato": "Antigo"})
            i += 1
            continue

        # Tenta Formato Novo (Bloco: Data -> Conteúdo -> Hora)
        if RE_DATA_BLOCO.match(linha_atual):
            data_str = linha_atual.replace("-", "/")
            bloco = []
            j = i + 1
            while j < len(linhas) and not RE_DATA_BLOCO.match(linhas[j]["texto"]):
                bloco.append(linhas[j]["texto"])
                j += 1
            
            log(f"\n[DEBUG] Bloco detectado - Pag {linhas[i]['pagina']} - Data: {data_str}")
            
            hora, temp = None, None
            for idx, item in enumerate(bloco):
                item_limpo = item.replace(",", ".")
                if RE_HORA_BLOCO.match(item_limpo):
                    hora = item_limpo
                    # Busca temperatura nas 3 linhas anteriores à hora
                    for k in range(idx-1, max(-1, idx-4), -1):
                        val = bloco[k].replace(",", ".")
                        if RE_NUM.match(val):
                            temp = float(val)
                            break
                    break
            
            if hora and temp is not None:
                dh = dt.datetime.strptime(f"{data_str} {hora}", "%d/%m/%Y %H:%M:%S")
                registros.append({"DataHora": dh, "TS": temp, "Formato": "Novo"})
                log(f"   -> Sucesso: {hora} | Temp: {temp}")
            i = j
        else:
            i += 1
    return registros

# ==========================================
# 5. EXECUÇÃO PRINCIPAL
# ==========================================
linhas = extrair_texto_pdf(PDF_PATH)
if not linhas:
    raise RuntimeError("Falha total na extração. Verifique se o PDF não está corrompido.")

dados = processar_dados(linhas)
df = pd.DataFrame(dados)
df["Temperatura_Apurada"] = df["TS"] # Simplificado para o novo formato
df = df[(df["DataHora"] >= DATA_INICIO) & (df["DataHora"] <= DATA_FIM)].sort_values("DataHora")

print(f"[2/4] Dados processados: {len(df)} registros encontrados.")

# ==========================================
# 6. GRÁFICO E EXCEL
# ==========================================
if not df.empty:
    fig, ax = plt.subplots(figsize=(12, 6))
    ax.plot(df["DataHora"], df["Temperatura_Apurada"], color="#6E3AB8", marker="o", markersize=3, label="Temp. Apurada")
    
    # Faixas e Labels
    ax.axhspan(0, 4, color="#08DB08", alpha=0.2)
    ax.axhspan(-25, -12, color="#4FADE4", alpha=0.4)
    
    x_mid = df["DataHora"].min() + (df["DataHora"].max() - df["DataHora"].min()) / 2
    ax.text(x_mid, 2, "Área refrigerada", ha="center", fontweight="bold", color="#0B6B0B")
    ax.text(x_mid, -18.5, "Área congelada", ha="center", fontweight="bold", color="#0B3E66")
    
    ax.set_title(f"Monitoramento: {df['DataHora'].min().strftime('%d/%m %H:%M')} a {df['DataHora'].max().strftime('%d/%m %H:%M')}")
    ax.yaxis.set_major_locator(ticker.MultipleLocator(5))
    ax.xaxis.set_major_formatter(DateFormatter("%d/%m %H:%M"))
    ax.grid(True, alpha=0.3)
    ax.legend()
    
    plt.tight_layout()
    fig.savefig(OUT_PNG, dpi=150)
    
    # Salvar Excel
    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Dados")
    
    wb = load_workbook(OUT_XLSX)
    ws = wb.create_sheet("Gráfico")
    ws.add_image(XLImage(OUT_PNG), "A1")
    ws["A35"], ws["A36"] = "Refrigerado: 0°C a 4°C", "Congelado: -12°C a -25°C"
    wb.save(OUT_XLSX)
    print(">>> Processo concluído com sucesso.")
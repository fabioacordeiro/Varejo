# -*- coding: utf-8 -*-
# Extrair temperaturas de PDF (Sascar) -> Excel + gráfico com faixas
# Fabio Cordeiro - versão 4.0
# Suporta dois formatos:
# 1) Formato antigo: data/hora e temperaturas na mesma linha
# 2) Formato novo: data em uma linha, bloco de dados abaixo e temperatura próxima da hora
# Inclui logs de depuração por bloco lido

import re
import os
import datetime as dt
from pathlib import Path

import pdfplumber
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
from matplotlib.dates import DateFormatter
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

# ==========================================
# 1. CONFIGURAÇÕES E CAMINHOS
# ==========================================
PDF_PATH = r"C:\\Fabio\\Desenvolvimento\\Varejo\\Temperatura\\Temp_sem_cabecalho.pdf"
OUT_XLSX = r"C:\\Fabio\\Desenvolvimento\\Varejo\\Temperatura\\Temperatura_Extraida_sem_cabecalho.xlsx"
OUT_PNG  = r"C:\\Fabio\Desenvolvimento\\Varejo\\Temperatura\\Grafico_Temperatura_arquivo_sem_cabecalho.png"

PLACA_VEICULO = "Transportadora Luiz Mauro Hergert - Placa: FQH1H55, FQS1A24"

DATA_INICIO = dt.datetime(2026, 3, 13, 0, 1)
DATA_FIM    = dt.datetime(2026, 3, 17, 23, 0)

APURADA_PRIORIDADE = ["TS", "T1", "T2", "T3"]

DEBUG = True

# ==========================================
# 2. REGEX
# ==========================================
RE_DATAHORA_ANTIGO = re.compile(r"(\d{2}/\d{2}/\d{4}\s+\d{2}:\d{2}:\d{2})")
RE_DATA_BLOCO = re.compile(r"^\d{2}[-/]\d{2}[-/]\d{4}$")
RE_HORA_BLOCO = re.compile(r"^\d{2}:\d{2}:\d{2}$")
RE_NUM = re.compile(r"^[\-–—−]?\d{1,2}(?:[\,\.]\d+)?$")

print(">>> Iniciando processamento linear...")

# ==========================================
# 3. FUNÇÕES AUXILIARES
# ==========================================
def log(msg):
    if DEBUG:
        print(msg)

def normalizar_numero(txt):
    txt = txt.strip().replace("—", "-").replace("–", "-").replace("−", "-")
    txt = re.sub(r"[^0-9\-\.,]+", "", txt)
    txt = txt.replace(",", ".")
    return txt

def escolher_temperatura_apurada(row):
    for c in APURADA_PRIORIDADE:
        v = row.get(c)
        if pd.notna(v):
            return v
    return None

def extrair_linhas_pdf(pdf_path):
    linhas = []

    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            text = page.extract_text(x_tolerance=2, y_tolerance=2) or ""
            if not text.strip():
                log(f"[DEBUG] Página {page_num}: sem texto extraído.")
                continue

            for raw_line in text.splitlines():
                line = " ".join(raw_line.split()).strip()
                if line:
                    linhas.append({
                        "pagina": page_num,
                        "texto": line
                    })

    return linhas

# ==========================================
# 4. PARSER FORMATO ANTIGO
# ==========================================
def parse_linha_formato_antigo(line):
    m = RE_DATAHORA_ANTIGO.search(line)
    if not m:
        return None

    dh_str = m.group(1)

    try:
        dh = dt.datetime.strptime(dh_str, "%d/%m/%Y %H:%M:%S")
    except:
        return None

    vals = {}

    for col in ["T1", "T2", "T3", "TS"]:
        mm = re.search(rf"\b{col}\b\s*([\-–—]?\d+[,\.]?\d*)", line, flags=re.IGNORECASE)
        if mm:
            txt = normalizar_numero(mm.group(1))
            try:
                vals[col] = float(txt)
            except:
                pass

    if not vals:
        candidates = re.findall(r"[\-–—]?\d{1,2}[,\.]?\d*", line)
        tail_raw = [candidates[i] for i in range(max(0, len(candidates)-4), len(candidates))]
        tail_float = []

        for t_str in tail_raw:
            t_clean = normalizar_numero(t_str)
            try:
                valor = float(t_clean)
                if -50 <= valor <= 50:
                    tail_float.append(valor)
            except:
                pass

        if len(tail_float) >= 1:
            mapping = {
                1: ["TS"],
                2: ["T1", "TS"],
                3: ["T1", "T2", "TS"],
                4: ["T1", "T2", "T3", "TS"]
            }
            cols_to_map = mapping.get(min(4, len(tail_float)))
            if cols_to_map:
                for c, v in zip(cols_to_map, tail_float[-len(cols_to_map):]):
                    vals.setdefault(c, v)

    if vals:
        return {
            "DataHora": dh,
            "T1": vals.get("T1"),
            "T2": vals.get("T2"),
            "T3": vals.get("T3"),
            "TS": vals.get("TS"),
            "Formato": "antigo"
        }

    return None

# ==========================================
# 5. PARSER FORMATO NOVO POR BLOCOS
# ==========================================
def parse_blocos_formato_novo(linhas_pdf):
    registros = []
    i = 0

    while i < len(linhas_pdf):
        linha = linhas_pdf[i]["texto"]

        if RE_DATA_BLOCO.match(linha):
            data_str = linha.replace("-", "/")
            pagina = linhas_pdf[i]["pagina"]

            bloco = []
            j = i + 1
            while j < len(linhas_pdf) and not RE_DATA_BLOCO.match(linhas_pdf[j]["texto"]):
                bloco.append(linhas_pdf[j]["texto"])
                j += 1

            log("")
            log(f"[DEBUG] --- BLOCO INICIADO ---")
            log(f"[DEBUG] Página: {pagina}")
            log(f"[DEBUG] Data encontrada: {data_str}")
            for item in bloco:
                log(f"[DEBUG]   {item}")

            hora_encontrada = None
            temperatura_encontrada = None
            hora_idx = None

            for idx_item, item in enumerate(bloco):
                item_limpo = item.replace("—", "-").replace("–", "-").replace("−", "-").replace(",", ".")
                if RE_HORA_BLOCO.match(item_limpo):
                    hora_encontrada = item_limpo
                    hora_idx = idx_item
                    break

            if hora_idx is not None:
                for k in range(hora_idx - 1, -1, -1):
                    item_limpo = bloco[k].replace("—", "-").replace("–", "-").replace("−", "-").replace(",", ".")
                    if RE_NUM.match(item_limpo):
                        try:
                            valor = float(item_limpo)
                            if -50 <= valor <= 50:
                                temperatura_encontrada = valor
                                break
                        except:
                            pass

            log(f"[DEBUG] Hora encontrada: {hora_encontrada}")
            log(f"[DEBUG] Temperatura encontrada: {temperatura_encontrada}")

            if data_str and hora_encontrada and temperatura_encontrada is not None:
                try:
                    dh = dt.datetime.strptime(f"{data_str} {hora_encontrada}", "%d/%m/%Y %H:%M:%S")
                    registros.append({
                        "DataHora": dh,
                        "T1": None,
                        "T2": None,
                        "T3": None,
                        "TS": temperatura_encontrada,
                        "Formato": "novo"
                    })
                    log(f"[DEBUG] Registro criado: {dh} | TS={temperatura_encontrada}")
                except Exception as e:
                    log(f"[DEBUG] Erro ao montar DataHora do bloco: {e}")

            log(f"[DEBUG] --- BLOCO FINALIZADO ---")
            i = j
        else:
            i += 1

    return registros

# ==========================================
# 6. EXTRAÇÃO DE DADOS
# ==========================================
if not os.path.exists(PDF_PATH):
    raise FileNotFoundError(f"PDF não encontrado: {PDF_PATH}")

print("[1/4] Lendo PDF...")
linhas_pdf = extrair_linhas_pdf(PDF_PATH)
print(f"   -> Linhas extraídas: {len(linhas_pdf)}")

registros = []

# 6.1 tenta formato antigo linha a linha
for item in linhas_pdf:
    rec = parse_linha_formato_antigo(item["texto"])
    if rec:
        registros.append(rec)
        log(f"[DEBUG] Formato antigo detectado: {rec['DataHora']} | {rec}")

# 6.2 tenta formato novo por blocos
registros_blocos = parse_blocos_formato_novo(linhas_pdf)
registros.extend(registros_blocos)

# Remove duplicados
df = pd.DataFrame(registros)

if df.empty:
    raise RuntimeError("Nenhum dado encontrado no PDF.")

if "DataHora" not in df.columns:
    raise KeyError(f"Coluna 'DataHora' não encontrada. Colunas disponíveis: {list(df.columns)}")

df = df.sort_values("DataHora").drop_duplicates(subset=["DataHora", "T1", "T2", "T3", "TS"]).reset_index(drop=True)

df["Temperatura_Apurada"] = df.apply(escolher_temperatura_apurada, axis=1)

if df["Temperatura_Apurada"].isna().all():
    raise RuntimeError("Os registros foram encontrados, mas nenhuma temperatura pôde ser apurada.")

print(f"[2/4] Filtrando dados ({DATA_INICIO} a {DATA_FIM})...")

mask = (df["DataHora"] >= DATA_INICIO) & (df["DataHora"] <= DATA_FIM)
df = df.loc[mask].copy()

if df.empty:
    raise RuntimeError(
        f"O filtro de data removeu todos os registros. "
        f"Período usado: {DATA_INICIO} até {DATA_FIM}"
    )
else:
    print(f"   -> Registros restantes: {len(df)}")

# ==========================================
# 7. GERAÇÃO DO GRÁFICO
# ==========================================
print("[3/4] Gerando gráfico com faixas e nova escala...")

fig, ax = plt.subplots(figsize=(12, 6))

ax.plot(
    df["DataHora"],
    df["Temperatura_Apurada"],
    linewidth=2,
    color="#6E3AB8",
    marker="o",
    markersize=3,
    label="Temp. Apurada"
)

ax.axhspan(0, 4, color="#08DB08", alpha=0.2, label="Refrigerado (0°C a 4°C)")
ax.axhspan(-25, -12, color="#4FADE4", alpha=0.4, label="Congelado (-12°C a -25°C)")

x0 = df["DataHora"].min()
x1 = df["DataHora"].max()
x_center = x0 + (x1 - x0) / 2

ax.text(
    x_center, 2, "Área refrigerada",
    ha="center", va="center",
    fontsize=11, fontweight="bold",
    color="#0B6B0B",
    bbox=dict(facecolor="white", edgecolor="none", alpha=0.6, pad=2)
)

ax.text(
    x_center, -18.5, "Área congelada",
    ha="center", va="center",
    fontsize=11, fontweight="bold",
    color="#0B3E66",
    bbox=dict(facecolor="white", edgecolor="none", alpha=0.6, pad=2)
)

ax.set_xlabel("Data/Hora")
ax.set_ylabel("Temperatura (°C)")

min_date_str = df["DataHora"].min().strftime("%d/%m %H:%M")
max_date_str = df["DataHora"].max().strftime("%d/%m %H:%M")
ax.set_title(f"Monitoramento de Temperatura ({min_date_str} até {max_date_str})")

ax.text(
    0.5, 0.92, f"{PLACA_VEICULO}",
    transform=ax.transAxes,
    ha="center", va="center",
    fontsize=12, fontweight="bold",
    color="black",
    bbox=dict(facecolor="white", edgecolor="none", alpha=0.7, pad=2)
)

ax.grid(True, linewidth=0.4, linestyle="--", alpha=0.6)
ax.yaxis.set_major_locator(ticker.MultipleLocator(5))
ax.xaxis.set_major_formatter(DateFormatter("%d/%m %H:%M"))
fig.autofmt_xdate()
ax.legend(loc="best", frameon=True, shadow=True)

plt.tight_layout()
fig.savefig(OUT_PNG, dpi=150)
plt.close(fig)

print(f"   -> Gráfico salvo em: {OUT_PNG}")

# ==========================================
# 8. SALVAR EXCEL
# ==========================================
print("[4/4] Salvando Excel final...")

with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
    df.to_excel(writer, index=False, sheet_name="Dados")

wb = load_workbook(OUT_XLSX)
ws_graf = wb.create_sheet("Gráfico")

if os.path.exists(OUT_PNG):
    img = XLImage(OUT_PNG)
    img.anchor = "A1"
    ws_graf.add_image(img)

ws_graf["A30"] = "Parâmetros utilizados:"
ws_graf["A31"] = f"Filtro Início: {DATA_INICIO}"
ws_graf["A32"] = f"Filtro Fim: {DATA_FIM}"
ws_graf["A34"] = "Legenda das Faixas:"
ws_graf["A35"] = "Refrigerado: 0°C a 4°C"
ws_graf["A36"] = "Congelado: -12°C a -25°C"

wb.save(OUT_XLSX)
print(f"   -> Excel salvo em: {OUT_XLSX}")

print("\n>>> Processo concluído com sucesso.")
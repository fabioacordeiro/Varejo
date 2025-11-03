# -*- coding: utf-8 -*-
# Extrair temperaturas de PDF (Sascar) -> Excel + gráfico com faixas
# Fabio Cordeiro - versão 1.0

import re
import os
import math
import datetime as dt
from pathlib import Path

import pdfplumber
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.dates import DateFormatter
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage

# =========================
# CONFIGURAÇÕES
# =========================
# Caminho do PDF (ajuste se necessário)
PDF_PATH = r"C:\\Fabio\Desenvolvimento\\Varejo\\Temperatura\\Temp.pdf"
# Saídas
OUT_XLSX = r"C:\\Fabio\Desenvolvimento\\Varejo\\Temperatura\\Temperatura_Extraida.xlsx"
OUT_PNG  = r"C:\\Fabio\Desenvolvimento\\Varejo\\Temperatura\\Grafico_Temperatura.png"

# Qual coluna priorizar como "temperatura apurada"
# ordem de preferência (se a anterior estiver vazia, usa a próxima)
APURADA_PRIORIDADE = ["TS", "T1", "T2", "T3"]

# =========================
# FUNÇÕES AUXILIARES
# =========================
# Ex.: 13/10/2025 10:55:05 (UTC-3)
RE_DATAHORA = re.compile(r"(\d{2}/\d{2}/\d{4}\s+\d{2}:\d{2}:\d{2})")

def _parse_float_cell(txt: str):
    """
    Converte a célula de temperatura, lidando com casos como:
      "-7", "--7", " -0", "N/A", "-" etc.
    Retorna float ou None.
    """
    if not txt:
        return None
    s = str(txt).strip().upper()
    if s in {"", "N/A", "NA"}:
        return None
    # remover duplicidade de sinais e caracteres "—", etc.
    s = s.replace("—", "-").replace("−", "-")
    s = re.sub(r"^-{2,}", "-", s)  # "--19" -> "-19"
    s = re.sub(r"[^0-9\-\.,]+", "", s)  # remove lixo, mantém números e sinais
    if s == "" or s == "-":
        return None
    # vírgula para ponto
    s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        m = re.search(r"-?\d+(?:\.\d+)?", s)
        return float(m.group(0)) if m else None

def _parse_datetime(br_str: str):
    # "13/10/2025 10:55:05" -> datetime
    return dt.datetime.strptime(br_str, "%d/%m/%Y %H:%M:%S")

def escolher_temperatura_apurada(row) -> float | None:
    for col in APURADA_PRIORIDADE:
        v = row.get(col)
        if v is not None and not (isinstance(v, float) and math.isnan(v)):
            return v
    return None

# =========================
# EXTRAÇÃO DO PDF
# =========================
def extrair_do_pdf(pdf_path: str) -> pd.DataFrame:
    """
    Percorre todas as páginas. Procura linhas que contenham uma data/hora e,
    na mesma linha, tenta identificar colunas T1, T2, T3, TS.
    Como o relatório é tabular mas pode vir "texto solto", usamos regex.
    """
    registros = []

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            # extrair texto com quebras de linha
            text = page.extract_text(x_tolerance=2, y_tolerance=2) or ""
            if not text.strip():
                continue

            # analisamos linha a linha
            for raw_line in text.splitlines():
                line = " ".join(raw_line.split())  # normaliza espaços
                # procurar data/hora
                m = RE_DATAHORA.search(line)
                if not m:
                    continue
                dh_str = m.group(1)
                try:
                    dh = _parse_datetime(dh_str)
                except Exception:
                    continue

                # Tenta capturar T1/T2/T3/TS procurando números perto dos rótulos.
                # Como o layout pode variar, usamos várias abordagens.
                # 1) Padrão com rótulos explícitos "T1", "T2", "T3", "TS"
                vals = {}

                # Procura pares tipo "T1 <valor>"
                for col in ["T1", "T2", "T3", "TS"]:
                    mm = re.search(rf"\b{col}\b\s*([\-–—]?\d+[,\.]?\d*)", line, flags=re.IGNORECASE)
                    if mm:
                        vals[col] = _parse_float_cell(mm.group(1))

                # 2) Se a linha não traz os rótulos, pode estar em colunas fixas.
                #    Como fallback, extrai a "última sequência" de números possíveis
                #    e tenta mapear posições para T1..TS quando houver 3-4 valores.
                if not vals:
                    # Pega TODOS os números com sinal que façam sentido
                    candidates = re.findall(r"[\-–—]?\d{1,2}[,\.]?\d*", line)
                    # heurística: temperaturas costumam estar no fim da linha
                    # guardamos os últimos 4
                    tail = [candidates[i] for i in range(max(0, len(candidates)-4), len(candidates))]
                    tail = [_parse_float_cell(t) for t in tail]
                    tail = [t for t in tail if t is not None]
                    if len(tail) >= 1:
                        # mapear: se 1 -> TS; se 2 -> T1, TS; se 3 -> T1,T2,TS; se 4 -> T1,T2,T3,TS
                        mapping_seq = {
                            1: ["TS"],
                            2: ["T1", "TS"],
                            3: ["T1", "T2", "TS"],
                            4: ["T1", "T2", "T3", "TS"],
                        }
                        cols = mapping_seq.get(min(4, len(tail)))
                        if cols:
                            for c, v in zip(cols, tail[-len(cols):]):
                                vals.setdefault(c, v)

                if not vals:
                    continue

                rec = {"DataHora": dh,
                       "T1": vals.get("T1"),
                       "T2": vals.get("T2"),
                       "T3": vals.get("T3"),
                       "TS": vals.get("TS")}
                registros.append(rec)

    df = pd.DataFrame(registros).sort_values("DataHora").reset_index(drop=True)
    # coluna temperatura apurada
    df["Temperatura_Apurada"] = df.apply(escolher_temperatura_apurada, axis=1)
    return df

# =========================
# GRÁFICO (PNG) + EXCEL
# =========================
def criar_grafico_png(df: pd.DataFrame, out_png: str):
    if df.empty:
        raise RuntimeError("Não há dados para plotar.")

    fig, ax = plt.subplots(figsize=(10, 4))
    ax.plot(df["DataHora"], df["Temperatura_Apurada"], linewidth=1.5)

    # Faixas:
    # ≤ -12°C (preenche de -40 até -12)
    ax.axhspan(-40, -12, alpha=0.2)
    # 0 a 5°C
    ax.axhspan(0, 5, alpha=0.2)

    ax.set_xlabel("Data/Hora")
    ax.set_ylabel("Temperatura (°C)")
    ax.set_title("Temperatura Apurada e Faixas")
    ax.grid(True, linewidth=0.4, linestyle="--", alpha=0.6)

    ax.xaxis.set_major_formatter(DateFormatter("%d/%m %H:%M"))
    fig.autofmt_xdate()

    plt.tight_layout()
    fig.savefig(out_png, dpi=150)
    plt.close(fig)

def salvar_excel_com_grafico(df: pd.DataFrame, out_xlsx: str, img_path: str):
    # salva os dados
    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Dados")

    # insere imagem do gráfico em uma aba "Gráfico"
    wb = load_workbook(out_xlsx)
    ws = wb.create_sheet("Gráfico")
    if Path(img_path).exists():
        img = XLImage(img_path)
        img.anchor = "A1"
        ws.add_image(img)
    # legenda das faixas
    ws["A30"] = "Faixas destacadas:"
    ws["A31"] = "≤ -12 °C (região azulada)"
    ws["A32"] = "0 a 5 °C (região azulada)"
    wb.save(out_xlsx)

# =========================
# MAIN
# =========================
def main():
    pdf_file = PDF_PATH
    if not Path(pdf_file).exists():
        raise FileNotFoundError(f"PDF não encontrado: {pdf_file}")

    print("[1/3] Extraindo dados do PDF...")
    df = extrair_do_pdf(pdf_file)

    if df.empty:
        raise RuntimeError("Nenhuma linha de temperatura foi reconhecida no PDF.")

    print("[2/3] Gerando gráfico...")
    criar_grafico_png(df, OUT_PNG)

    print("[3/3] Salvando Excel com dados e gráfico...")
    salvar_excel_com_grafico(df, OUT_XLSX, OUT_PNG)

    # resumo
    print(f"Linhas extraídas: {len(df)}")
    print(f"Excel: {OUT_XLSX}")
    print(f"Gráfico (PNG): {OUT_PNG}")

if __name__ == "__main__":
    main()
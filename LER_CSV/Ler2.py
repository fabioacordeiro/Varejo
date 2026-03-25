import os
import csv
import threading
import traceback
import queue
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from openpyxl import Workbook


class ConversorCSVExcelApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Conversor de CSV para Excel - Fabio A Cordeiro")
        self.root.geometry("1180x760")
        self.root.minsize(1080, 700)
        self.root.configure(bg="#0b0b0b")

        self.fila_ui = queue.Queue()
        self.processando = False

        self.pasta_var = tk.StringVar()
        self.saida_var = tk.StringVar()
        self.modo_var = tk.StringVar(value="pasta")
        self.status_var = tk.StringVar(value="Pronto para iniciar.")
        self.csvs_selecionados = []

        self.criar_estilo()
        self.criar_layout()
        self.atualizar_fila_ui()

    # =========================
    # ESTILO
    # =========================
    def criar_estilo(self):
        style = ttk.Style()
        try:
            style.theme_use("clam")
        except Exception:
            pass

        cor_fundo = "#0b0b0b"
        cor_card = "#151515"
        cor_card2 = "#1c1c1c"
        cor_texto = "#f2f2f2"
        cor_texto_sec = "#b8b8b8"
        cor_laranja = "#ff8c00"
        cor_verde = "#16c47f"
        cor_borda = "#2a2a2a"

        style.configure("TFrame", background=cor_fundo)
        style.configure("Card.TFrame", background=cor_card, relief="flat")
        style.configure("InnerCard.TFrame", background=cor_card2, relief="flat")

        style.configure(
            "Titulo.TLabel",
            background=cor_fundo,
            foreground=cor_laranja,
            font=("Segoe UI", 22, "bold")
        )
        style.configure(
            "Subtitulo.TLabel",
            background=cor_fundo,
            foreground=cor_texto_sec,
            font=("Segoe UI", 10)
        )
        style.configure(
            "Cabecalho.TLabel",
            background=cor_card,
            foreground=cor_texto,
            font=("Segoe UI", 11, "bold")
        )
        style.configure(
            "Campo.TLabel",
            background=cor_card,
            foreground=cor_texto,
            font=("Segoe UI", 10)
        )
        style.configure(
            "Info.TLabel",
            background=cor_card,
            foreground=cor_texto_sec,
            font=("Segoe UI", 9)
        )
        style.configure(
            "Status.TLabel",
            background=cor_fundo,
            foreground=cor_verde,
            font=("Segoe UI", 10, "bold")
        )

        style.configure(
            "Primario.TButton",
            font=("Segoe UI", 10, "bold"),
            padding=(14, 10),
            background=cor_laranja,
            foreground="#ffffff",
            borderwidth=0,
            focuscolor="none"
        )
        style.map(
            "Primario.TButton",
            background=[("active", "#ff9f1a"), ("disabled", "#5b5b5b")],
            foreground=[("disabled", "#d0d0d0")]
        )

        style.configure(
            "Secundario.TButton",
            font=("Segoe UI", 10, "bold"),
            padding=(14, 10),
            background="#262626",
            foreground="#ffffff",
            borderwidth=0,
            focuscolor="none"
        )
        style.map(
            "Secundario.TButton",
            background=[("active", "#333333"), ("disabled", "#5b5b5b")],
            foreground=[("disabled", "#d0d0d0")]
        )

        style.configure(
            "TNotebook",
            background=cor_fundo,
            borderwidth=0,
            tabmargins=[0, 0, 0, 0]
        )
        style.configure(
            "TNotebook.Tab",
            background="#202020",
            foreground="#f0f0f0",
            padding=(18, 10),
            font=("Segoe UI", 10, "bold"),
            borderwidth=0
        )
        style.map(
            "TNotebook.Tab",
            background=[("selected", cor_laranja), ("active", "#2c2c2c")],
            foreground=[("selected", "#ffffff")]
        )

        style.configure(
            "Horizontal.TProgressbar",
            troughcolor="#1b1b1b",
            background=cor_laranja,
            lightcolor=cor_laranja,
            darkcolor=cor_laranja,
            bordercolor="#1b1b1b"
        )

        self.cores = {
            "fundo": cor_fundo,
            "card": cor_card,
            "card2": cor_card2,
            "texto": cor_texto,
            "texto_sec": cor_texto_sec,
            "laranja": cor_laranja,
            "borda": cor_borda
        }

    # =========================
    # LAYOUT
    # =========================
    def criar_layout(self):
        main = ttk.Frame(self.root, padding=18)
        main.pack(fill="both", expand=True)

        topo = ttk.Frame(main)
        topo.pack(fill="x", pady=(0, 12))

        ttk.Label(
            topo,
            text="Conversor de CSV para Excel - Fabio A Cordeiro",
            style="Titulo.TLabel"
        ).pack(anchor="w")

        ttk.Label(
            topo,
            text="Leitura rápida de CSVs da pasta/subpastas ou arquivos específicos, com geração de Excel consolidado e aba de erros.",
            style="Subtitulo.TLabel"
        ).pack(anchor="w", pady=(4, 0))

        card_config = ttk.Frame(main, style="Card.TFrame", padding=18)
        card_config.pack(fill="x", pady=(0, 12))

        ttk.Label(card_config, text="⚙ Configurações", style="Cabecalho.TLabel").grid(
            row=0, column=0, sticky="w", columnspan=3, pady=(0, 12)
        )

        ttk.Label(card_config, text="Arquivo Excel de saída", style="Campo.TLabel").grid(
            row=1, column=0, sticky="w"
        )

        self.entry_saida = tk.Entry(
            card_config,
            textvariable=self.saida_var,
            bg="#1b1b1b",
            fg="#ffffff",
            insertbackground="#ffffff",
            relief="flat",
            font=("Segoe UI", 10)
        )
        self.entry_saida.grid(row=2, column=0, sticky="ew", padx=(0, 10), ipady=10, pady=(6, 0))

        self.btn_saida = ttk.Button(
            card_config,
            text="💾 Salvar como",
            style="Secundario.TButton",
            command=self.escolher_saida
        )
        self.btn_saida.grid(row=2, column=1, sticky="ew", pady=(6, 0))

        self.btn_processar = ttk.Button(
            card_config,
            text="▶ Gerar Excel",
            style="Primario.TButton",
            command=self.iniciar_processamento
        )
        self.btn_processar.grid(row=2, column=2, sticky="ew", pady=(6, 0))

        ttk.Label(
            card_config,
            text="Dica: para arquivos grandes, o Excel é gerado em fluxo para consumir menos memória.",
            style="Info.TLabel"
        ).grid(row=3, column=0, columnspan=3, sticky="w", pady=(10, 0))

        card_config.columnconfigure(0, weight=1)

        notebook = ttk.Notebook(main)
        notebook.pack(fill="x", pady=(0, 12))

        # Aba 1 - pasta
        self.aba_pasta = ttk.Frame(notebook, style="Card.TFrame", padding=18)
        notebook.add(self.aba_pasta, text="📁 Pasta e subpastas")

        ttk.Label(self.aba_pasta, text="Selecionar pasta de origem", style="Cabecalho.TLabel").grid(
            row=0, column=0, sticky="w", columnspan=2, pady=(0, 10)
        )

        self.entry_pasta = tk.Entry(
            self.aba_pasta,
            textvariable=self.pasta_var,
            bg="#1b1b1b",
            fg="#ffffff",
            insertbackground="#ffffff",
            relief="flat",
            font=("Segoe UI", 10)
        )
        self.entry_pasta.grid(row=1, column=0, sticky="ew", padx=(0, 10), ipady=10)

        self.btn_pasta = ttk.Button(
            self.aba_pasta,
            text="📂 Selecionar pasta",
            style="Secundario.TButton",
            command=self.escolher_pasta
        )
        self.btn_pasta.grid(row=1, column=1, sticky="ew")

        ttk.Label(
            self.aba_pasta,
            text="Este modo localiza automaticamente todos os arquivos .csv da pasta e de todas as subpastas.",
            style="Info.TLabel"
        ).grid(row=2, column=0, columnspan=2, sticky="w", pady=(10, 0))

        self.aba_pasta.columnconfigure(0, weight=1)

        # Aba 2 - arquivos específicos
        self.aba_arquivos = ttk.Frame(notebook, style="Card.TFrame", padding=18)
        notebook.add(self.aba_arquivos, text="🗂 Escolher CSVs")

        ttk.Label(self.aba_arquivos, text="Selecionar arquivos CSV específicos", style="Cabecalho.TLabel").pack(
            anchor="w", pady=(0, 10)
        )

        botoes_arqs = ttk.Frame(self.aba_arquivos)
        botoes_arqs.pack(fill="x")

        self.btn_add_csvs = ttk.Button(
            botoes_arqs,
            text="➕ Adicionar CSVs",
            style="Secundario.TButton",
            command=self.adicionar_csvs
        )
        self.btn_add_csvs.pack(side="left")

        self.btn_limpar_csvs = ttk.Button(
            botoes_arqs,
            text="🧹 Limpar lista",
            style="Secundario.TButton",
            command=self.limpar_csvs
        )
        self.btn_limpar_csvs.pack(side="left", padx=(10, 0))

        self.lista_csvs = tk.Listbox(
            self.aba_arquivos,
            bg="#111111",
            fg="#f2f2f2",
            selectbackground="#ff8c00",
            selectforeground="#ffffff",
            relief="flat",
            font=("Consolas", 10),
            height=10
        )
        self.lista_csvs.pack(fill="x", pady=(12, 0))

        ttk.Label(
            self.aba_arquivos,
            text="Use esta aba quando quiser processar apenas alguns CSVs, sem varrer a pasta inteira.",
            style="Info.TLabel"
        ).pack(anchor="w", pady=(10, 0))

        # Barra de progresso
        progresso_card = ttk.Frame(main, style="Card.TFrame", padding=18)
        progresso_card.pack(fill="x", pady=(0, 12))

        ttk.Label(progresso_card, text="📊 Progresso", style="Cabecalho.TLabel").pack(anchor="w", pady=(0, 10))

        self.progress = ttk.Progressbar(
            progresso_card,
            mode="determinate",
            style="Horizontal.TProgressbar"
        )
        self.progress.pack(fill="x")

        ttk.Label(
            progresso_card,
            textvariable=self.status_var,
            style="Status.TLabel"
        ).pack(anchor="w", pady=(10, 0))

        # Log
        log_card = ttk.Frame(main, style="Card.TFrame", padding=18)
        log_card.pack(fill="both", expand=True)

        topo_log = ttk.Frame(log_card)
        topo_log.pack(fill="x", pady=(0, 10))

        ttk.Label(topo_log, text="📝 Log de processamento", style="Cabecalho.TLabel").pack(side="left")

        self.btn_limpar_log = ttk.Button(
            topo_log,
            text="🧽 Limpar log",
            style="Secundario.TButton",
            command=self.limpar_log
        )
        self.btn_limpar_log.pack(side="right")

        frame_texto = ttk.Frame(log_card, style="InnerCard.TFrame")
        frame_texto.pack(fill="both", expand=True)

        self.txt_log = tk.Text(
            frame_texto,
            bg="#101010",
            fg="#eaeaea",
            insertbackground="#ffffff",
            relief="flat",
            wrap="word",
            font=("Consolas", 10),
            padx=10,
            pady=10
        )
        self.txt_log.pack(side="left", fill="both", expand=True)

        scrollbar = tk.Scrollbar(frame_texto, command=self.txt_log.yview, bg="#202020")
        scrollbar.pack(side="right", fill="y")
        self.txt_log.config(yscrollcommand=scrollbar.set)

        self.log("Sistema iniciado com sucesso.")

    # =========================
    # UI HELPERS
    # =========================
    def log(self, msg):
        self.txt_log.insert("end", f"{msg}\n")
        self.txt_log.see("end")

    def limpar_log(self):
        self.txt_log.delete("1.0", "end")
        self.log("Log limpo.")

    def set_status(self, msg):
        self.status_var.set(msg)

    def set_progress(self, valor):
        self.progress["value"] = valor

    def escolher_pasta(self):
        pasta = filedialog.askdirectory(title="Selecione a pasta com os CSVs")
        if pasta:
            self.pasta_var.set(pasta)
            if not self.saida_var.get().strip():
                nome = os.path.join(pasta, "consolidado_csv.xlsx")
                self.saida_var.set(nome)
            self.log(f"Pasta selecionada: {pasta}")

    def escolher_saida(self):
        arquivo = filedialog.asksaveasfilename(
            title="Salvar Excel como",
            defaultextension=".xlsx",
            filetypes=[("Arquivo Excel", "*.xlsx")]
        )
        if arquivo:
            self.saida_var.set(arquivo)
            self.log(f"Arquivo de saída: {arquivo}")

    def adicionar_csvs(self):
        arquivos = filedialog.askopenfilenames(
            title="Selecione os arquivos CSV",
            filetypes=[("Arquivos CSV", "*.csv"), ("Todos os arquivos", "*.*")]
        )
        if arquivos:
            novos = 0
            for arq in arquivos:
                if arq not in self.csvs_selecionados:
                    self.csvs_selecionados.append(arq)
                    self.lista_csvs.insert("end", arq)
                    novos += 1
            self.log(f"{novos} arquivo(s) CSV adicionados à lista.")

            if arquivos and not self.saida_var.get().strip():
                pasta_base = os.path.dirname(arquivos[0])
                self.saida_var.set(os.path.join(pasta_base, "consolidado_csv.xlsx"))

    def limpar_csvs(self):
        self.csvs_selecionados.clear()
        self.lista_csvs.delete(0, "end")
        self.log("Lista de CSVs específicos foi limpa.")

    def bloquear_controles(self, bloquear=True):
        estado = "disabled" if bloquear else "normal"
        for w in [
            self.btn_saida,
            self.btn_processar,
            self.btn_pasta,
            self.btn_add_csvs,
            self.btn_limpar_csvs,
            self.btn_limpar_log,
        ]:
            w.config(state=estado)

    def atualizar_fila_ui(self):
        try:
            while True:
                item = self.fila_ui.get_nowait()
                acao = item[0]

                if acao == "log":
                    self.log(item[1])
                elif acao == "status":
                    self.set_status(item[1])
                elif acao == "progress":
                    self.set_progress(item[1])
                elif acao == "fim_ok":
                    self.finalizar_sucesso(item[1], item[2], item[3], item[4])
                elif acao == "fim_erro":
                    self.finalizar_erro(item[1])
        except queue.Empty:
            pass

        self.root.after(120, self.atualizar_fila_ui)

    # =========================
    # PROCESSAMENTO
    # =========================
    def iniciar_processamento(self):
        if self.processando:
            messagebox.showwarning("Atenção", "Já existe um processamento em andamento.")
            return

        saida = self.saida_var.get().strip()
        if not saida:
            messagebox.showwarning("Atenção", "Informe o arquivo Excel de saída.")
            return

        if not saida.lower().endswith(".xlsx"):
            saida += ".xlsx"
            self.saida_var.set(saida)

        arquivos_csv = self.obter_arquivos_entrada()
        if not arquivos_csv:
            messagebox.showwarning("Atenção", "Nenhum arquivo CSV foi localizado.")
            return

        self.processando = True
        self.bloquear_controles(True)
        self.set_progress(0)
        self.set_status("Iniciando processamento...")
        self.log("=" * 70)
        self.log("Processamento iniciado.")
        self.log(f"Total de arquivos CSV: {len(arquivos_csv)}")

        thread = threading.Thread(
            target=self.processar_arquivos,
            args=(arquivos_csv, saida),
            daemon=True
        )
        thread.start()

    def obter_arquivos_entrada(self):
        # Prioriza lista específica, se houver arquivos nela
        if self.csvs_selecionados:
            return list(self.csvs_selecionados)

        pasta = self.pasta_var.get().strip()
        if not pasta:
            return []

        arquivos = []
        for raiz, _, nomes in os.walk(pasta):
            for nome in nomes:
                if nome.lower().endswith(".csv"):
                    arquivos.append(os.path.join(raiz, nome))
        return arquivos

    def detectar_encoding_e_delimitador(self, caminho):
        encodings = ["utf-8-sig", "utf-8", "cp1252", "latin1"]
        delimitadores = [";", ",", "\t", "|"]

        for enc in encodings:
            try:
                with open(caminho, "r", encoding=enc, newline="") as f:
                    primeira = f.readline()
                    segunda = f.readline()
                    amostra = (primeira or "") + (segunda or "")

                if not amostra.strip():
                    return enc, ";"

                contagens = {d: amostra.count(d) for d in delimitadores}
                delimitador = max(contagens, key=contagens.get)

                if contagens[delimitador] == 0:
                    delimitador = ";"

                return enc, delimitador
            except Exception:
                continue

        raise Exception("Não foi possível detectar encoding/delimitador.")

    def ler_cabecalho_csv(self, caminho):
        enc, delim = self.detectar_encoding_e_delimitador(caminho)

        with open(caminho, "r", encoding=enc, newline="") as f:
            leitor = csv.reader(f, delimiter=delim)
            for linha in leitor:
                if linha and any(str(x).strip() for x in linha):
                    cab = [str(c).strip() for c in linha]
                    return enc, delim, cab

        raise Exception("Arquivo vazio ou sem cabeçalho válido.")

    def iterar_linhas_csv_dict(self, caminho, enc, delim):
        with open(caminho, "r", encoding=enc, newline="") as f:
            leitor = csv.DictReader(f, delimiter=delim)
            for row in leitor:
                if row is None:
                    continue
                yield row

    def processar_arquivos(self, arquivos_csv, saida):
        erros = []
        total_linhas = 0
        metadados = []
        cabecalho_final = []
        cabecalho_set = set()

        try:
            # PASSO 1: descobrir cabeçalhos
            self.fila_ui.put(("status", "Analisando estrutura dos arquivos CSV..."))

            total = len(arquivos_csv)
            for i, caminho in enumerate(arquivos_csv, start=1):
                try:
                    enc, delim, cab = self.ler_cabecalho_csv(caminho)

                    if not cab:
                        raise Exception("Cabeçalho vazio.")

                    for col in cab:
                        if col not in cabecalho_set:
                            cabecalho_set.add(col)
                            cabecalho_final.append(col)

                    metadados.append({
                        "arquivo": caminho,
                        "encoding": enc,
                        "delim": delim,
                        "cabecalho": cab
                    })

                    self.fila_ui.put((
                        "log",
                        f"[OK] Estrutura lida: {os.path.basename(caminho)} | encoding={enc} | separador={repr(delim)} | colunas={len(cab)}"
                    ))

                except Exception as e:
                    msg = str(e)
                    erros.append([caminho, "Leitura de cabeçalho", msg])
                    self.fila_ui.put((
                        "log",
                        f"[ERRO] {os.path.basename(caminho)} | {msg}"
                    ))

                progresso = (i / max(total, 1)) * 20
                self.fila_ui.put(("progress", progresso))

            if not metadados:
                raise Exception("Nenhum CSV válido foi encontrado para consolidar.")

            colunas_extras = ["ARQUIVO_ORIGEM", "CAMINHO_ORIGEM"]
            for col in colunas_extras:
                if col not in cabecalho_set:
                    cabecalho_final.append(col)

            # PASSO 2: gerar excel em fluxo
            self.fila_ui.put(("status", "Gerando Excel consolidado em modo otimizado..."))

            wb = Workbook(write_only=True)
            ws = wb.create_sheet(title="Consolidado")
            ws_erros = wb.create_sheet(title="Erros")

            ws.append(cabecalho_final)
            ws_erros.append(["Arquivo", "Etapa", "Erro"])

            total_validos = len(metadados)

            for i, meta in enumerate(metadados, start=1):
                caminho = meta["arquivo"]
                enc = meta["encoding"]
                delim = meta["delim"]

                try:
                    linhas_arquivo = 0

                    for row in self.iterar_linhas_csv_dict(caminho, enc, delim):
                        linha_saida = []
                        for col in cabecalho_final:
                            if col == "ARQUIVO_ORIGEM":
                                linha_saida.append(os.path.basename(caminho))
                            elif col == "CAMINHO_ORIGEM":
                                linha_saida.append(caminho)
                            else:
                                valor = row.get(col, "")
                                if valor is None:
                                    valor = ""
                                linha_saida.append(valor)

                        ws.append(linha_saida)
                        total_linhas += 1
                        linhas_arquivo += 1

                        if total_linhas % 5000 == 0:
                            self.fila_ui.put((
                                "status",
                                f"Escrevendo Excel... {total_linhas:,} linhas consolidadas."
                            ))

                    self.fila_ui.put((
                        "log",
                        f"[OK] Arquivo consolidado: {os.path.basename(caminho)} | linhas={linhas_arquivo}"
                    ))

                except Exception as e:
                    msg = str(e)
                    erros.append([caminho, "Leitura de dados", msg])
                    self.fila_ui.put((
                        "log",
                        f"[ERRO] Falha ao consolidar {os.path.basename(caminho)} | {msg}"
                    ))

                progresso = 20 + (i / max(total_validos, 1)) * 70
                self.fila_ui.put(("progress", progresso))

            # grava erros
            for err in erros:
                ws_erros.append(err)

            self.fila_ui.put(("status", "Salvando arquivo Excel..."))
            self.fila_ui.put(("progress", 95))

            wb.save(saida)

            self.fila_ui.put(("progress", 100))
            self.fila_ui.put(("fim_ok", saida, len(arquivos_csv), total_linhas, len(erros)))

        except Exception as e:
            detalhe = f"{str(e)}\n\n{traceback.format_exc()}"
            self.fila_ui.put(("fim_erro", detalhe))

    # =========================
    # FINALIZAÇÃO
    # =========================
    def finalizar_sucesso(self, saida, total_arquivos, total_linhas, total_erros):
        self.processando = False
        self.bloquear_controles(False)
        self.set_status("Processamento concluído com sucesso.")
        self.log("-" * 70)
        self.log("PROCESSAMENTO FINALIZADO")
        self.log(f"Arquivo gerado: {saida}")
        self.log(f"Arquivos localizados: {total_arquivos}")
        self.log(f"Linhas consolidadas: {total_linhas:,}")
        self.log(f"Registros de erro: {total_erros}")
        self.log("-" * 70)

        messagebox.showinfo(
            "Sucesso",
            "Processamento concluído com sucesso.\n\n"
            f"Arquivo gerado:\n{saida}\n\n"
            f"Arquivos localizados: {total_arquivos}\n"
            f"Linhas consolidadas: {total_linhas:,}\n"
            f"Registros de erro: {total_erros}"
        )

    def finalizar_erro(self, erro):
        self.processando = False
        self.bloquear_controles(False)
        self.set_status("Erro durante o processamento.")
        self.log("-" * 70)
        self.log("ERRO NO PROCESSAMENTO")
        self.log(erro)
        self.log("-" * 70)

        messagebox.showerror("Erro", erro)


if __name__ == "__main__":
    root = tk.Tk()
    app = ConversorCSVExcelApp(root)
    root.mainloop()
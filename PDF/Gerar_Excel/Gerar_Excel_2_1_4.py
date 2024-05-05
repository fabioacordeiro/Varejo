#pip install pypdf2 #necessário instalar
import os # Leitura arquivos, pastas e diretórios
from pathlib import Path
import chardet #pip install chardet #para ver qual o tipo de dados do csv
import csv
from openpyxl.cell import Cell, MergedCell
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.packaging.relationship import RelationshipList
from openpyxl.workbook.child import _WorkbookChild
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook import Workbook
from openpyxl.cell import Cell  # Aqui eu corrijo a tipagem de Cell
from openpyxl.worksheet.worksheet import Worksheet
import time #Tempo de execução
import pandas as pd
import tabula #Manipulação de Tabelas 
import PyPDF2 # import PdfMerger, PdfReader, PdfWriter #Manipular pdf
import PyPDF2.pagerange
from pathlib import Path #Manipular pastas, diretórios e arquivos
#from openpyxl import Workbook, worksheet
from openpyxl.utils.dataframe import dataframe_to_rows

# PyPDF2 para manipular arquivos PDF (PdfMerger)
# PyPDF2 é uma biblioteca de manipulação de arquivos PDF feita em Python puro,
# gratuita e de código aberto. Ela é capaz de ler, manipular, escrever e unir
# dados de arquivos PDF, assim como adicionar anotações, transformar páginas.
#------------------------------------------
#Pegando o tempo de execução do código
tempo_inicial = time.time()
seconds = 0
hour = 0
minutes = 0
print("--- %s segundos ---" % (time.time() - tempo_inicial))
nomeprograma = 'Gerador de Excel'
print(50*'_')
print(f'{nomeprograma:^50}')
print(50*'_')
#-------------Fim Variação tempo -------------------------------
#Incluindo o Try e except que força a execução do código e detalhe
# de erros de execução
lista_pdf = []
lista_pdf2 = []
try:
    PASTA_RAIZ = Path(__file__).parent
    PASTA_ORIGINAIS = PASTA_RAIZ / 'Arquivo_pdf'
    CRIADO1 = PASTA_RAIZ / 'Arquivo_csv'
    CRIADO2 = PASTA_RAIZ / 'Arquivo_Excel'
    PASTA_FINAL = PASTA_ORIGINAIS/'Teste.pdf'
    print(' ----- Início ---- ') 
    print('Caminhos onde serão lidos e gravados os arquivos')    
    print(f'Caminho PDFS_ORIGINAIS:\n',(PASTA_ORIGINAIS))
    print(f'Caminho Raiz:\n',(PASTA_RAIZ))
    print(f'Caminho arquivo csv criado:\n',(CRIADO1))
    print(f'Caminho arquivo XlS criado:\n',(CRIADO2))
    print(f'Caminho PASTA_FINAL:\n',(PASTA_FINAL))
    print('Passo 1')
    reader = PyPDF2.PdfReader(PASTA_ORIGINAIS/'Teste.pdf')
    page = len(reader.pages)
    print(f"Quantidade de páginas do pdf:{page}")
    for p in reader.pages:
        #print(p.extract_text())
        lista_pdf.append(p.extract_text())
        
    print('Passo 2')
    print(f"------- Criando arquivo CSV -------")
    with open(CRIADO1/'Teste_1.csv', 'w', newline='\n') as d:
            writer = csv.writer(d)
            writer.writerow(lista_pdf)
    print('Passo 3')  
    # Step 2: Read CSV File in Binary Mode
    with open(CRIADO1/'Teste_1.csv','rb') as f:
        data = f.read()
    
    # Step 3: Detect Encoding using chardet Library
    encoding_result = chardet.detect(data)
    # Step 4: Retrieve Encoding Information
    encoding = encoding_result['encoding']
    # Step 5: Print Detected Encoding Information
    #print("Detected Encoding:", encoding)
    print(f"------- Transformando CSV em Dataframe -------")
    N3 = pd.read_csv(CRIADO1/"Teste_1.csv",encoding='latin-1')
    N4 = []
    for linha in N3:
        N4.append(linha)
    DF0 =pd.DataFrame(N4) 
    print('Passo 4')
    print(f'Transformando Dataframe em Excel (páginas em linhas)') 
    DF0.to_excel(CRIADO2/'Teste_1.xlsx')    
    print('Passo 5')
    print(f'---- Ler Excel e manipular extração ----')
    wb = load_workbook(CRIADO2/"Teste_1.xlsx")
    ws = wb['Sheet1']
    page0_Sheet1 = ws
    ws1 = wb.create_sheet("Fabio")
    page1_Fabio = ws1
    for rows in page0_Sheet1.iter_rows(min_col=2,max_col=2, min_row=1, max_row=73):
        for cell in rows:
            page1_Fabio.append([cell.value.lstrip])
    wb.save(CRIADO2/"Teste_1.xlsx")       

    print('-'*65)
    print('Passei ponto 5')
    print('Fim')
except ZeroDivisionError:
    print('Dividiu por zero')
except NameError:
    print('Algum nome não está definido')
except(TypeError, IndexError) as error:
    print('TypeError + IndexError')
    print('MSG:', error)
    print('MSG:', error.__class__.__name__)
except Exception:
    print('ERRO DESCONHECIDO')
else:
    print('Não deu erro')
finally:
    print('Fechar arquivo')
print('Fim')
#----------------------------------------------------
#Imprimindo o tempo inicial menos o final, ou seja, tempo total
seconds = (time.time() - tempo_inicial)
hour = seconds // 3600
minutes = seconds // 60
seconds %= 60
print("--- %s segundos ---" % (time.time() - tempo_inicial))
print("--- %s segundos ---" % ("%d:%02d:%02d" % (hour, minutes, seconds) ))
